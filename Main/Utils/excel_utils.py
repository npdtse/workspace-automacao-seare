from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side
import sys
import logging
import os

file_handler = logging.FileHandler("Log/excel_utils.log", "a")
file_handler.setLevel(logging.ERROR)

stream_handler = logging.StreamHandler()

logging.basicConfig(
    level=logging.INFO,
    format="%(levelname)s - %(asctime)s - %(message)s",
    handlers=[file_handler, stream_handler]
)

class ExcelUtils:

    def __init__(self, caminho_excel):
        self.caminho_excel = caminho_excel

    def extrair_nsei(self):
        
        try:
            workbook = load_workbook(self.caminho_excel) #Carrega o arquivo excel especificado no parâmetro
            worksheet = workbook.worksheets[0] #Define a primeira aba da planilha como a principal

            ultima_linha = worksheet.max_row #Define a ultima linha preenchida
            nsei_list = []

            for linha in range(3, ultima_linha + 1): #Itera sobre a planilha começando na linha 3 até a ultima linha
                nsei = worksheet[f"B{linha}"].value #Pega o valor do número SEI da célula e linha atual

                if not nsei : logging.warning(f"Na linha {linha}, não foi encontrado o número SEI"); continue #Se a célula estiver vazia, passa para a próxima iteração

                nsei_list.append(nsei) #Salva o número sei na lista

            workbook.close() #Fecha a planilha
            logging.info(f"Extração concluída com sucesso. Total de NSEIs extraídos: {len(nsei_list)}")    
            return nsei_list
        
        except Exception as e:
            logging.error("Não foi possível extrair o número SEI: {e}", exc_info=True)
            sys.exit(1)

    def salvar_nsei_na_planilha(self, nsei_list, excel_historico):

        try:
            workbook = load_workbook(excel_historico) #Carrega o arquivo excel especificado no parâmetro
            worksheet = workbook.worksheets[0] #Define a primeira aba da planilha como a principal

            for linha, nsei in enumerate(nsei_list, start=2): #Fazemos a iteração da lista de Números SEI usando o enumerate como um contador, ou seja, o contador inicia na linha 2
                worksheet[f"A{linha}"].value = nsei #Coloca o número SEI na célula A da linha correspondente, Ex: A2 = nsei(atual)

            workbook.save(excel_historico) #Salva a planilha
            workbook.close() #Fecha a planilha
            logging.info(f"NSEIs salvos com sucesso na planilha: {excel_historico}")
            
        except Exception as e:
            logging.error(f"Erro ao salvar o NSEIs: {e}", exc_info=True)
            sys.exit(1)

    def atualizar_titulos(self, excel_historico ,titulos_list):

        try:
            logging.info(f"Atualizando títulos na planilha: {excel_historico}")
            workbook = load_workbook(excel_historico) #Carrega o arquivo excel especificado no parâmetro
            worksheet = workbook.worksheets[0] #Define a primeira aba da planilha como a principal

            #Define a borda padrão de cada célula
            borda_padrao = Border(
                top=Side(border_style="thin", color="000000"),
                left=Side(border_style="thin", color="000000"),
                right=Side(border_style="thin", color="000000"),
                bottom=Side(border_style="thin", color="000000"),
            )

            titulos_alterados = []
            linha_atual = 2

            for titulo_novo in titulos_list:
                celula_nsei = worksheet[f"A{linha_atual}"] #Define qual a coluna e a linha da célula dos números SEI 
                celula_titulo = worksheet[f"B{linha_atual}"] #Define qual a coluna e a linha da célula dos títulos

                titulo_antigo = celula_titulo.value #Cria uma variavel titulo_antigo e define o valor dela como o valor atual da célula título

                celula_titulo.fill = PatternFill("solid", start_color="FFFFFF") #Define a cor da célula do título como Branca
                celula_nsei.border = borda_padrao #Aplica a borda padrão na célula dos números SEI
                celula_titulo.border = borda_padrao #Aplica a borda padrão na célula dos títulos

                if titulo_antigo != titulo_novo: #Verifica se o titulo_antigo é diferente do título que está vindo da lista de títulos
                    celula_titulo.value = titulo_novo #Se a condição for verdadeira, adiciona o novo titulo na célula e linha atual dos títulos
                    if titulo_novo == "Processo Restrito" or titulo_novo == "Aguarde...":
                        celula_titulo.fill = PatternFill("solid", start_color="D37F79") #Define a cor da célula como verde para saber que houve alteração
                    else:
                        celula_titulo.fill = PatternFill("solid", start_color="8ED973")   
                        
                    titulos_alterados.append(titulo_novo) #Salva o novo titulo na lista

                linha_atual = linha_atual + 1
            workbook.save(excel_historico) #Salva a planilha
            workbook.close() #Fecha o Arquivo
            logging.info("Atualização de títulos concluída com sucesso.")
            return titulos_alterados
        except Exception as e:
            logging.error(f"Erro ao atualizar títulos: {e}", exc_info=True)
            sys.exit(1)
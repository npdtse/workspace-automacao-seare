from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side
import warnings
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import tkinter as tk
from tkinter import messagebox
import time
import os
from tkinter.filedialog import askopenfile, askopenfilename

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

def obter_credenciais():
    global login, senha
    login = entry_login.get()
    senha = entry_senha.get()

    if login and senha:
        root.destroy()
    else:
        messagebox.showwarning("Atenção", "Preencha o login e a senha")

def extrair_nsei_e_salvar_na_planilha(caminho_excel):

    #Caminho do arquivo excel
    excel_path = caminho_excel

    #Carregar a planilha
    workBook = load_workbook(excel_path)
    workSheet = workBook.worksheets[0]

    ultima_linha = workSheet.max_row

    nsei_list = []

    for linha in range(3, ultima_linha + 1):
        nsei = workSheet[f"B{linha}"].value

        if not nsei : print("Campo Número SEI está vazio"); continue

        nsei_list.append(nsei)

    workBook.close

    save_nsei(nsei_list, caminho_excel)

    print(nsei_list)

    return nsei_list

def save_nsei(lista_nsei, caminho_excel_historico):

    try:
        workbook = load_workbook(caminho_excel_historico)
        workSheet = workbook.worksheets[0]

        primeira_linha = 2
        
        linha_atual = 2

        for nsei in lista_nsei:
            workSheet[f"A{linha_atual}"] = nsei
            linha_atual = linha_atual + 1

        workbook.save(caminho_excel_historico)
        workbook.close()

    except Exception as e:
        print(f"Erro ao carregar o arquivo: {e}")

def logar_no_sei_e_buscar_titulos(nsei_list, site_url):

    navegador = logar_sei(site_url)

    lista_titulo_documento = []
    contador = 0
    for current_sei in nsei_list:
        numero_sei = current_sei
        
        campo_pesquisa = navegador.find_element(By.ID, "txtPesquisaRapida")
        campo_pesquisa.send_keys(numero_sei)
        campo_pesquisa.send_keys(Keys.RETURN)

        time.sleep(2)

        iframe = navegador.find_element(By.CSS_SELECTOR, "#divInfraAreaGlobal #divInfraAreaTela #divInfraAreaTelaD #divIframeArvore #ifrArvore")
        navegador.switch_to.frame(iframe)

        #Executa o script JavaScript dentro do iframe
        titulo = navegador.execute_script("""
            let iframeDoc = document;
            
            let docTitles = iframeDoc.querySelectorAll("body #container #content #frmArvore #divArvore div.infraArvore a.infraArvoreNo span");

            let signedDocTitles = Array.from(docTitles).filter(title => {
                let style = title.getAttribute("style");
                let id = title.getAttribute("id");
                return !(style && style.includes("color: rgb(204, 158, 128)")) && !(id && id.includes("PASTA"));
            }).map(title => title.textContent); 

            let lastSignedDocTitle = signedDocTitles[signedDocTitles.length - 1];
            
            return lastSignedDocTitle;
        """)

        lista_titulo_documento.append(titulo)
        contador = contador + 1
        print(f"Contador: {contador}: {numero_sei} : {titulo}")
        
        navegador.switch_to.default_content()

    time.sleep(2)
    return lista_titulo_documento

def atualizar_titulo_planilha_historico(caminho_excel_historico,lista_nsei,lista_titulos):

    try:
        workbook = load_workbook(caminho_excel_historico)
        worksheet = workbook.worksheets[0]

        titulos_alterados = []

        borda_padrao = Border(
            top=Side(border_style="thin", color="000000"),
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000"),
        )

        linha_atual = 2

        if worksheet[f"A{linha_atual}"].value is None:
            save_nsei(lista_nsei,caminho_excel_historico)

        if worksheet[f"B{linha_atual}"].value is None:
            save_title(lista_titulos, caminho_excel_historico)

        else:
            
            for titulo_novo in lista_titulos:
                celula_titulo = worksheet[f"B{linha_atual}"]
                celula_nsei = worksheet[f"A{linha_atual}"]

                titulo_antigo = celula_titulo.value

                celula_titulo.fill = PatternFill("solid", start_color= "FFFFFF") #Cor padrão excel
                celula_titulo.border = borda_padrao
                celula_nsei.border = borda_padrao

                if titulo_antigo != titulo_novo:
                    celula_titulo.value = titulo_novo
                    celula_titulo.fill = PatternFill("solid",start_color="8ED973") #Cor Verde
                    titulos_alterados.append(titulo_novo)
                linha_atual = linha_atual + 1

            workbook.save(caminho_excel_historico)
            workbook.close()

        print(titulos_alterados)
        return titulos_alterados
           

    except Exception as e:
        print(f"Erro ao carregar o arquivo: {e}")

def save_title(lista_titulos, caminho_excel_historico):

    try:
        workBook = load_workbook(caminho_excel_historico)
        workSheet = workBook.worksheets[0]

        ultima_linha = 2

        while workSheet[f"B{ultima_linha}"].value is not None:
            ultima_linha = ultima_linha + 1

        for titulo in lista_titulos:
            workSheet[f"B{ultima_linha}"] = titulo
            ultima_linha = ultima_linha + 1

        workBook.save(caminho_excel_historico)  
        workBook.close() 

    except Exception as e:
        print(f"Erro ao carregar o arquivo: {e}")   

def selecionar_arquivo():

    try:
        # caminho_arquivo = askopenfile()
        nome_arquivo = askopenfilename()
        if nome_arquivo:
          file_path = os.path.basename(nome_arquivo)
          print(file_path)
        return None
    except Exception:
        print("Arquivo Incorreto.") 

def main():



    #Dados para Login no SEI
    # login = "marcos.fernandes"
    # senha = "y@1KFMy77v78"
    site_url = 'https://sip.tse.jus.br/sip/login.php?sigla_orgao_sistema=TSE&sigla_sistema=SEI'

    #Caminho da planilha do computador de casa
    file_path_home = r"C:\Users\Marcos\OneDrive - TRIBUNAL SUPERIOR ELEITORAL\Documentos\Excel\Estudos Pessoais\Cronograma Execução PCA 2025\Cronograma de Execução do PCA 2025 - versão 16.2 Atualizada.xlsx"
    file_path_excel_history_home = r"C:\Users\Marcos\OneDrive - TRIBUNAL SUPERIOR ELEITORAL\Documentos\Excel\Estudos Pessoais\Cronograma Execução PCA 2025\Historico Varredura SEI.xlsx"

    #caminho da planilha do computador do Tribunal
    file_path_tse = r"D:\Users\marcos.fernandes\OneDrive - TRIBUNAL SUPERIOR ELEITORAL\Documentos\Excel\Estudos Pessoais\Cronograma Execução PCA 2025\Cronograma de Execução do PCA 2025 - versão 16.2 Atualizada.xlsx"
    file_path_excel_history_tse = r"D:\Users\marcos.fernandes\OneDrive - TRIBUNAL SUPERIOR ELEITORAL\Documentos\Excel\Estudos Pessoais\Cronograma Execução PCA 2025\Historico Varredura SEI.xlsx"



    #Dados para TESTE
    # nsei_list = ["2022.00.000008912-0", "2024.00.000000972-1"]
    # titulos_list = ["TesteQuaseDefinitivo1", "TesteQuaseDefinitivo123123431"]
    # novos_dados = {"2022.00.000008912-0" : "Teste4", "2024.00.000000972-1": "teste6" }


    #Funções para execução do Script
    # tempo_inicial = time.time()

    # nsei_list = extrair_nsei_e_salvar_na_planilha(file_path_tse)

    # titulos_list = logar_no_sei_e_buscar_titulos(nsei_list, site_url)

    # atualizar_titulo_planilha_historico(file_path_excel_history_tse,nsei_list, titulos_list)

    # tempo_final = time.time()

    selecionar_arquivo()

    # tempo_execução(tempo_inicial, tempo_final)

def tempo_execução(tempo_inicial, tempo_final):
    tempo_total = tempo_final - tempo_inicial
    minutos = int(tempo_total // 60)
    segundos = int(tempo_total % 60)

    print(f"Tempo de execução: {minutos} min {segundos} seg")

def logar_sei(site_url):
    # site_url = 'https://sip.tse.jus.br/sip/login.php?sigla_orgao_sistema=TSE&sigla_sistema=SEI'
    navegador = webdriver.Chrome()
    navegador.get(site_url)
    navegador.maximize_window()

    try:
        navegador.find_element(By.ID, "txtUsuario").send_keys(login)
        navegador.find_element(By.ID, "pwdSenha").send_keys(senha)
        navegador.find_element(By.ID, "sbmAcessar").click()

        time.sleep(2)

        if "controlador.php?" in navegador.current_url:
            return navegador

    except Exception as e:
        messagebox.showinfo("Usuario ou Senha incorreta:",  f"Login: {login} Senha: {senha}")

if __name__ == "__main__":
    main()

####################################################INICIO TELA DE CREDENCIAL DO SEI##########################################################################

# #Caminho favIcon TSE
# file_path_fav_icon_tse = r'D:\Users\marcos.fernandes\OneDrive - TRIBUNAL SUPERIOR ELEITORAL\Documentos\Excel\Estudos Pessoais\Projeto SEARE\tse-icon.png'

# #Criação da tela
# root = tk.Tk()
# root.iconphoto(False, tk.PhotoImage(file=file_path_fav_icon_tse))
# root.title("Login SEI")
# root.geometry("300x200")
# root.resizable(False, False)

# #Criando widgets
# tk.Label(root, text="Usuário:").pack(pady=5)
# entry_login = tk.Entry(root)
# entry_login.pack(pady=5)

# tk.Label(root, text="Senha:").pack(pady=5)
# entry_senha = tk.Entry(root, show="*")
# entry_senha.pack(pady=5)

# btn_entrar = tk.Button(root, text="Entrar", command=obter_credenciais)
# btn_entrar.pack(pady=20)

# root.mainloop()





  